"""Utility functions for parsing Excel files."""


def parse_excel_l12a(file_path: str) -> list[tuple]:
    """Parse the Excel file and return list of (L1, Agents) pairs from sheet 'L12A'.

    This function locates the header columns by searching for header cells with
    text 'L1' and 'Agents' (preferably bold). It then iterates rows below the
    header and returns pairs where both cells are non-empty and not bold.

    Args:
        file_path: Path to the Excel file.

    Returns:
        A list of tuples, where each tuple is (value_in_L1_col, value_in_Agents_col).

    Raises:
        ImportError: if `openpyxl` is not installed.
        ValueError: if the workbook does not contain a sheet named 'L12A' or the
                    required headers cannot be located.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise ImportError("openpyxl is required to parse Excel files") from e

    # Do not use read_only mode because we need access to cell.font.bold
    wb = load_workbook(filename=file_path, data_only=True)
    try:
        if 'L12A' not in wb.sheetnames:
            raise ValueError("Sheet 'L12A' not found in workbook")

        ws = wb['L12A']

        col_l1 = None
        col_agents = None
        header_row = None

        # Scan the first few rows to find header cells by bold formatting
        max_scan_row = min(10, ws.max_row or 10)
        for r in range(1, max_scan_row + 1):
            for c in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if isinstance(val, str) and val.strip() == 'L1' and getattr(cell.font, 'bold', False):
                    col_l1 = c
                    header_row = r
                if isinstance(val, str) and val.strip().lower() == 'agents' and getattr(cell.font, 'bold', False):
                    col_agents = c
                    header_row = r if header_row is None else header_row
                if col_l1 and col_agents:
                    break
            if col_l1 and col_agents:
                break

        # Fallback: find by text ignoring bold if necessary
        if col_l1 is None or col_agents is None:
            for r in range(1, max_scan_row + 1):
                for c in range(1, (ws.max_column or 1) + 1):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    if isinstance(val, str) and val.strip() == 'L1' and col_l1 is None:
                        col_l1 = c
                        header_row = r
                    if isinstance(val, str) and val.strip().lower() == 'agents' and col_agents is None:
                        col_agents = c
                        header_row = r if header_row is None else header_row
                if col_l1 and col_agents:
                    break

        if col_l1 is None or col_agents is None:
            raise ValueError("Could not locate header columns 'L1' and 'Agents' in sheet 'L12A'")

        pairs: list[tuple] = []

        for r in range((header_row or 1) + 1, (ws.max_row or 1) + 1):
            cell_b = ws.cell(row=r, column=col_l1)
            cell_c = ws.cell(row=r, column=col_agents)
            v_b = cell_b.value
            v_c = cell_c.value

            # Skip empty or whitespace-only values
            if v_b is None or (isinstance(v_b, str) and v_b.strip() == ''):
                continue
            if v_c is None or (isinstance(v_c, str) and v_c.strip() == ''):
                continue

            # Ensure the values themselves are not bold (headers are bold)
            bold_b = getattr(cell_b.font, 'bold', False)
            bold_c = getattr(cell_c.font, 'bold', False)
            if bold_b or bold_c:
                continue

            pairs.append((v_b, v_c))

        return pairs
    finally:
        wb.close()


def parse_excel_a2t(file_path: str) -> list[tuple]:
    """Parse the Excel file and return list of (Agents, Tools) pairs from sheet 'A2T'.

    This function locates the header columns by searching for header cells with
    text 'Agents' and 'Tools' (preferably bold). It then iterates rows below the
    header and returns pairs where both cells are non-empty and not bold.

    Args:
        file_path: Path to the Excel file.

    Returns:
        A list of tuples, where each tuple is (value_in_Agents_col, value_in_Tools_col).

    Raises:
        ImportError: if `openpyxl` is not installed.
        ValueError: if the workbook does not contain a sheet named 'A2T' or the
                    required headers cannot be located.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise ImportError("openpyxl is required to parse Excel files") from e

    # Do not use read_only mode because we need access to cell.font.bold
    wb = load_workbook(filename=file_path, data_only=True)
    try:
        if 'A2T' not in wb.sheetnames:
            raise ValueError("Sheet 'A2T' not found in workbook")

        ws = wb['A2T']

        col_agents = None
        col_tools = None
        header_row = None

        # Scan the first few rows to find header cells by bold formatting
        max_scan_row = min(10, ws.max_row or 10)
        for r in range(1, max_scan_row + 1):
            for c in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if isinstance(val, str) and val.strip().lower() == 'agents' and getattr(cell.font, 'bold', False):
                    col_agents = c
                    header_row = r
                if isinstance(val, str) and val.strip().lower() == 'tools' and getattr(cell.font, 'bold', False):
                    col_tools = c
                    header_row = r if header_row is None else header_row
                if col_agents and col_tools:
                    break
            if col_agents and col_tools:
                break

        # Fallback: find by text ignoring bold if necessary
        if col_agents is None or col_tools is None:
            for r in range(1, max_scan_row + 1):
                for c in range(1, (ws.max_column or 1) + 1):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    if isinstance(val, str) and val.strip().lower() == 'agents' and col_agents is None:
                        col_agents = c
                        header_row = r
                    if isinstance(val, str) and val.strip().lower() == 'tools' and col_tools is None:
                        col_tools = c
                        header_row = r if header_row is None else header_row
                if col_agents and col_tools:
                    break

        if col_agents is None or col_tools is None:
            raise ValueError("Could not locate header columns 'Agents' and 'Tools' in sheet 'A2T'")

        pairs: list[tuple] = []

        for r in range((header_row or 1) + 1, (ws.max_row or 1) + 1):
            cell_a = ws.cell(row=r, column=col_agents)
            cell_t = ws.cell(row=r, column=col_tools)
            v_a = cell_a.value
            v_t = cell_t.value

            # Skip empty or whitespace-only values
            if v_a is None or (isinstance(v_a, str) and v_a.strip() == ''):
                continue
            if v_t is None or (isinstance(v_t, str) and v_t.strip() == ''):
                continue

            # Ensure the values themselves are not bold (headers are bold)
            bold_a = getattr(cell_a.font, 'bold', False)
            bold_t = getattr(cell_t.font, 'bold', False)
            if bold_a or bold_t:
                continue

            pairs.append((v_a, v_t))

        return pairs
    finally:
        wb.close()


def parse_excel_components(file_path: str) -> tuple[list[str], list[int]]:
    """Parse the Excel file and return two lists from the 'Components' sheet.

    This function locates header columns 'Component_Type', 'Component_Name', and
    'Progress' in the 'Components' sheet. It then iterates through rows and returns:
    - list_1: names from Component_Name column
    - list_2: integers based on Component_Type (3 for L1, 2 for Agent, 1 for Tool)

    Args:
        file_path: Path to the Excel file.

    Returns:
        A tuple of (list_names, list_types) where:
            - list_names: list of component names (strings)
            - list_types: list of integers (3 for L1, 2 for Agent, 1 for Tool)

    Raises:
        ImportError: if `openpyxl` is not installed.
        ValueError: if the workbook does not contain a sheet named 'Components' or the
                    required headers cannot be located.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise ImportError("openpyxl is required to parse Excel files") from e

    # Do not use read_only mode because we need access to cell.font.bold
    wb = load_workbook(filename=file_path, data_only=True)
    try:
        if 'Components' not in wb.sheetnames:
            raise ValueError("Sheet 'Components' not found in workbook")

        ws = wb['Components']

        col_type = None
        col_name = None
        col_progress = None
        header_row = None

        # Scan the first few rows to find header cells by bold formatting
        max_scan_row = min(10, ws.max_row or 10)
        for r in range(1, max_scan_row + 1):
            for c in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if isinstance(val, str) and val.strip().lower() == 'component_type' and getattr(cell.font, 'bold', False):
                    col_type = c
                    header_row = r
                if isinstance(val, str) and val.strip().lower() == 'component_name' and getattr(cell.font, 'bold', False):
                    col_name = c
                    header_row = r if header_row is None else header_row
                if isinstance(val, str) and val.strip().lower() == 'progress' and getattr(cell.font, 'bold', False):
                    col_progress = c
                    header_row = r if header_row is None else header_row
                if col_type and col_name and col_progress:
                    break
            if col_type and col_name and col_progress:
                break

        # Fallback: find by text ignoring bold if necessary
        if col_type is None or col_name is None or col_progress is None:
            for r in range(1, max_scan_row + 1):
                for c in range(1, (ws.max_column or 1) + 1):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    if isinstance(val, str) and val.strip().lower() == 'component_type' and col_type is None:
                        col_type = c
                        header_row = r
                    if isinstance(val, str) and val.strip().lower() == 'component_name' and col_name is None:
                        col_name = c
                        header_row = r if header_row is None else header_row
                    if isinstance(val, str) and val.strip().lower() == 'progress' and col_progress is None:
                        col_progress = c
                        header_row = r if header_row is None else header_row
                if col_type and col_name and col_progress:
                    break

        if col_type is None or col_name is None or col_progress is None:
            raise ValueError("Could not locate header columns 'Component_Type', 'Component_Name', and 'Progress' in sheet 'Components'")

        list_names: list[str] = []
        list_types: list[float] = []

        for r in range((header_row or 1) + 1, (ws.max_row or 1) + 1):
            cell_type = ws.cell(row=r, column=col_type)
            cell_name = ws.cell(row=r, column=col_name)
            v_type = cell_type.value
            v_name = cell_name.value

            # Skip if name is empty or whitespace-only
            if v_name is None or (isinstance(v_name, str) and v_name.strip() == ''):
                continue

            # Skip if type is empty
            if v_type is None:
                continue

            # Ensure the values themselves are not bold (headers are bold)
            bold_type = getattr(cell_type.font, 'bold', False)
            bold_name = getattr(cell_name.font, 'bold', False)
            if bold_type or bold_name:
                continue

            # Add name to list
            list_names.append(v_name)

            # Map component type to integer
            type_str = str(v_type).strip().lower()
            if type_str == 'l1':
                list_types.append(1.5)
            elif type_str == 'agent':
                list_types.append(1)
            elif type_str == 'tool':
                list_types.append(.5)
            else:
                # Skip unknown types or default to 0
                list_types.append(0)

        return (list_names, list_types)
    finally:
        wb.close()

